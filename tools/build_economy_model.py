"""Builds design/economy_model.xlsx — the Phase 0 30-day cash-flow model.

Exit criterion (roadmap Phase 0): three solvent-but-DIFFERENT paths.
Income taxonomy maps to background gates:
  Spot sales        - anyone
  Negotiated (rep)  - Old School legacy contracts
  Timed (info)      - IT Nephew market-timing edge
  Repair/Flip       - Mechanic wrench income
  Eggs & favors     - ambient trickle
Numbers here are the tuning source of truth; they get transcribed to /data JSON.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

BLUE = Font(name="Arial", size=10, color="0000FF")
BLACK = Font(name="Arial", size=10)
GREEN = Font(name="Arial", size=10, color="008000")
BOLD = Font(name="Arial", size=10, bold=True)
H1 = Font(name="Arial", size=13, bold=True)
H2 = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HDRFILL = PatternFill("solid", start_color="4A3F2A")
YELLOW = PatternFill("solid", start_color="FFFF00")
MONEY = '$#,##0;($#,##0);"-"'
MONEY2 = '$#,##0.00;($#,##0.00);"-"'
PCT = "0.0%"
THIN = Border(bottom=Side(style="thin", color="CCCCCC"))

wb = Workbook()

# ---------------- Assumptions ----------------
A = wb.active
A.title = "Assumptions"
A.column_dimensions["A"].width = 34
for c in "BCDEFGH":
    A.column_dimensions[c].width = 13

def put(ws, ref, val, font=BLACK, fmt=None, fill=None):
    ws[ref] = val
    ws[ref].font = font
    if fmt:
        ws[ref].number_format = fmt
    if fill:
        ws[ref].fill = fill

put(A, "A1", "DIRT MONEY — ECONOMY MODEL v1 (Phase 0)", H1)
put(A, "A2", "All blue cells are tunable inputs. Background sheets reference them. Source: docs/DIRT_MONEY_ECONOMY_SPEC.md, 2026-07-03.", Font(name="Arial", size=9, italic=True))

rows = [
    ("GLOBAL", None, None),
    ("Starting cash ($)", 1200, MONEY),
    ("Starting debt ($)", 8000, MONEY),
    ("Debt daily interest", 0.004, PCT),
    ("Travel fuel per trip ($)", 8, MONEY),
    ("Chickens (head)", 12, "0"),
    ("Eggs per chicken per day", 0.8, "0.0"),
    ("Egg price ($/egg)", 0.40, MONEY2),
    ("Feed per chicken per day ($)", 0.08, MONEY2),
    ("CROPS (per field cycle)", None, None),
]
r = 4
for label, val, fmt in rows:
    if val is None:
        put(A, f"A{r}", label, H2); A[f"A{r}"].fill = HDRFILL
    else:
        put(A, f"A{r}", label, BLACK)
        put(A, f"B{r}", val, BLUE, fmt)
    r += 1
G = {"cash": "$B$5", "debt": "$B$6", "int": "$B$7", "fuel": "$B$8",
     "chick": "$B$9", "epc": "$B$10", "eprice": "$B$11", "feed": "$B$12"}

crop_hdr = ["Crop", "Plant cost $", "Plant days", "Grow days", "Harvest cost $", "Harvest days", "Yield (units)", "Price $/unit"]
for i, h in enumerate(crop_hdr):
    put(A, f"{get_column_letter(1+i)}14", h, H2); A[f"{get_column_letter(1+i)}14"].fill = HDRFILL
crops = [("Corn", 340, 2, 10, 180, 2, 300, 4.20),
         ("Soybeans", 220, 2, 9, 140, 1, 200, 10.50),
         ("Hay (per cut)", 90, 1, 6, 80, 1, 100, 5.00)]
for i, row in enumerate(crops):
    for j, v in enumerate(row):
        put(A, f"{get_column_letter(1+j)}{15+i}", v, BLUE if j else BLACK,
            MONEY2 if j == 7 else (MONEY if j in (1, 4) else "0"))
C = {"corn": 15, "soy": 16, "hay": 17}  # rows

bg_rows = [
    ("BACKGROUND MODIFIERS", None, None),
    ("Old School: legacy premium (× spot)", 1.35, "0.00x"),
    ("Old School: legacy contract units", 450, "0"),
    ("Old School: Hollis hauling favor ($)", 120, MONEY),
    ("IT Nephew: labor cost multiplier", 1.5, "0.00x"),
    ("IT Nephew: market-timing edge (× spot)", 1.15, "0.00x"),
    ("IT Nephew: egg contract premium (× price)", 1.2, "0.00x"),
    ("Mechanic: repair job pay ($)", 250, MONEY),
    ("Mechanic: repair job parts cost ($)", 60, MONEY),
    ("Mechanic: neighbor repair pay ($)", 150, MONEY),
    ("Mechanic: flip - salvage buy ($)", 400, MONEY),
    ("Mechanic: flip - restoration parts ($)", 250, MONEY),
    ("Mechanic: flip - sale price ($)", 1400, MONEY),
]
r = 19
for label, val, fmt in bg_rows:
    if val is None:
        put(A, f"A{r}", label, H2); A[f"A{r}"].fill = HDRFILL
    else:
        put(A, f"A{r}", label)
        put(A, f"B{r}", val, BLUE, fmt)
    r += 1
M = {"legacy_x": "$B$20", "legacy_u": "$B$21", "favor": "$B$22", "it_labor": "$B$23",
     "it_edge": "$B$24", "it_egg": "$B$25", "rj_pay": "$B$26", "rj_parts": "$B$27",
     "nb_pay": "$B$28", "flip_buy": "$B$29", "flip_parts": "$B$30", "flip_sale": "$B$31"}
put(A, "A33", "Demo win condition: END CASH > 0 and net worth improved. Full loan payoff is multi-season (Phase 3) scope.", Font(name="Arial", size=9, italic=True))
A["A33"].fill = YELLOW

AS = "Assumptions!"
def ar(key):  # global/mod ref
    return AS + (G.get(key) or M[key])
def cr(crop, col):  # crop table ref: col B..H -> cost..price
    return f"{AS}${col}${C[crop]}"

# ---------------- Background sheets ----------------
COLS = ["Day", "Weekday", "Event", "Spot sales", "Negotiated (rep)", "Timed (info)",
        "Repair & flip", "Eggs & favors", "Costs", "Net", "Cash", "Debt", "Net worth"]
DAY0, NDAYS = 6, 30  # data starts row 6

def sheet_scaffold(name, strategy):
    ws = wb.create_sheet(name)
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 9
    ws.column_dimensions["C"].width = 44
    for c in range(4, 14):
        ws.column_dimensions[get_column_letter(c)].width = 13
    put(ws, "A1", f"{name.upper()} — 30-DAY CASH FLOW", H1)
    put(ws, "A2", strategy, Font(name="Arial", size=9, italic=True))
    ws.merge_cells("A2:M2")
    put(ws, "A3", "Income category gates: Negotiated = reputation-locked · Timed = info-locked · Repair & flip = wrench-locked", Font(name="Arial", size=8, italic=True, color="777777"))
    for i, h in enumerate(COLS):
        cell = f"{get_column_letter(1+i)}5"
        put(ws, cell, h, H2); ws[cell].fill = HDRFILL
    for d in range(1, NDAYS + 1):
        r = DAY0 + d - 1
        put(ws, f"A{r}", d, BLACK, "0")
        put(ws, f"B{r}", f'=CHOOSE(MOD(A{r}-1,7)+1,"Mon","Tue","Wed","Thu","Fri","Sat","Sun")', BLACK)
        for col in "DEFGHI":
            put(ws, f"{col}{r}", 0, BLACK, MONEY)
        put(ws, f"H{r}", f"=({ar('chick')}*{ar('epc')}*{ar('eprice')})-({ar('chick')}*{ar('feed')})", GREEN, MONEY)
        put(ws, f"J{r}", f"=SUM(D{r}:H{r})-I{r}", BLACK, MONEY)
        prev_cash = ar("cash") if d == 1 else f"K{r-1}"
        prev_debt = ar("debt") if d == 1 else f"L{r-1}"
        put(ws, f"K{r}", f"={prev_cash}+J{r}", BLACK, MONEY)
        put(ws, f"L{r}", f"={prev_debt}*(1+{ar('int')})", BLACK, MONEY)
        put(ws, f"M{r}", f"=K{r}-L{r}", BLACK, MONEY)
        ws[f"C{r}"].font = BLACK
    return ws

def ev(ws, day, text, **cells):
    r = DAY0 + day - 1
    existing = ws[f"C{r}"].value
    ws[f"C{r}"] = f"{existing}; {text}" if existing else text
    for col, formula in cells.items():
        cur = ws[f"{col}{r}"].value
        add = formula.lstrip("=")
        ws[f"{col}{r}"] = f"={add}" if cur in (0, None) else f"{ws[f'{col}{r}'].value}+{add}"
        ws[f"{col}{r}"].font = GREEN
        ws[f"{col}{r}"].number_format = MONEY

def summary(ws):
    r0, r1 = DAY0, DAY0 + NDAYS - 1
    r = r1 + 2
    put(ws, f"A{r}", "RUN SUMMARY", H2); ws[f"A{r}"].fill = HDRFILL
    items = [
        ("Total spot sales", f"=SUM(D{r0}:D{r1})"),
        ("Total negotiated (rep)", f"=SUM(E{r0}:E{r1})"),
        ("Total timed (info)", f"=SUM(F{r0}:F{r1})"),
        ("Total repair & flip", f"=SUM(G{r0}:G{r1})"),
        ("Total eggs & favors", f"=SUM(H{r0}:H{r1})"),
        ("Total costs", f"=SUM(I{r0}:I{r1})"),
        ("End cash", f"=K{r1}"),
        ("End debt", f"=L{r1}"),
        ("Net worth delta", f"=M{r1}-({ar('cash')}-{ar('debt')})"),
        ("Top income source", f'=INDEX({{"Spot sales";"Negotiated (rep)";"Timed (info)";"Repair & flip";"Eggs & favors"}},MATCH(MAX(B{r+1}:B{r+5}),B{r+1}:B{r+5},0))'),
    ]
    for i, (label, formula) in enumerate(items):
        put(ws, f"A{r+1+i}", label, BOLD)
        put(ws, f"B{r+1+i}", formula, BLACK, MONEY if i < 9 else None)
    return r + 1  # first summary data row

# --- Old School Farmer: reputation is the business model ---
os_ = sheet_scaffold("Old School", "Corn on both big fields + hay; the storm-help unlock converts most of the corn to the legacy contract at a premium rate. Wins on relationships, not information.")
ev(os_, 2, "Plant corn — North field", I=f"={cr('corn','B')}+{ar('fuel')}")
ev(os_, 3, "Plant corn — South field", I=f"={cr('corn','B')}+{ar('fuel')}")
ev(os_, 4, "Plant hay — East field", I=f"={cr('hay','B')}+{ar('fuel')}")
ev(os_, 5, "Haul Hollis's hay (favor)", H=f"={ar('favor')}", I=f"={ar('fuel')}")
ev(os_, 8, "STORM — help Hollis sandbag (loses the day; unlocks legacy contract)")
ev(os_, 10, "Hay cut 1 — East", D=f"={cr('hay','G')}*{cr('hay','H')}", I=f"={cr('hay','E')}")
ev(os_, 14, "Harvest corn — North", I=f"={cr('corn','E')}")
ev(os_, 15, "Harvest corn — South", I=f"={cr('corn','E')}")
ev(os_, 16, "Deliver legacy contract corn (the Vann family rate)", E=f"={ar('legacy_u')}*{cr('corn','H')}*{ar('legacy_x')}", I=f"={ar('fuel')}")
ev(os_, 17, "Sell remaining corn at spot", D=f"=(2*{cr('corn','G')}-{ar('legacy_u')})*{cr('corn','H')}", I=f"={ar('fuel')}")
ev(os_, 17, "Hay cut 2 — East", D=f"={cr('hay','G')}*{cr('hay','H')}", I=f"={cr('hay','E')}")
ev(os_, 24, "Hay cut 3 — East", D=f"={cr('hay','G')}*{cr('hay','H')}", I=f"={cr('hay','E')}")
os_sum = summary(os_)

# --- IT Nephew: information is the business model ---
it = sheet_scaffold("IT Nephew", "Higher labor costs force crop selection by the numbers: soybeans + corn, all sales timed off the forecast for the arbitrage edge. Egg contract renegotiated on day 1 with data. Wins on information, not relationships.")
ev(it, 1, "Audit Marge's egg contract — premium locked in (data check)")
for d in range(1, NDAYS + 1):  # egg premium overrides base egg formula
    r = DAY0 + d - 1
    it[f"H{r}"] = f"=({ar('chick')}*{ar('epc')}*{ar('eprice')}*{ar('it_egg')})-({ar('chick')}*{ar('feed')}*IF(A{r}>=20,0,1))"
    it[f"H{r}"].font = GREEN
    it[f"H{r}"].number_format = MONEY
ev(it, 2, "Plant corn — North (hired labor)", I=f"={cr('corn','B')}*{ar('it_labor')}+{ar('fuel')}")
ev(it, 3, "Plant soybeans — South (hired labor)", I=f"={cr('soy','B')}*{ar('it_labor')}+{ar('fuel')}")
ev(it, 12, "Harvest soybeans — South", I=f"={cr('soy','E')}*{ar('it_labor')}")
ev(it, 15, "Sell soybeans on forecast spike", F=f"={cr('soy','G')}*{cr('soy','H')}*{ar('it_edge')}", I=f"={ar('fuel')}")
ev(it, 14, "Harvest corn — North", I=f"={cr('corn','E')}*{ar('it_labor')}")
ev(it, 17, "Sell corn on forecast spike", F=f"={cr('corn','G')}*{cr('corn','H')}*{ar('it_edge')}", I=f"={ar('fuel')}")
ev(it, 20, "Coop automation online — feed cost eliminated (perk)")
it_sum = summary(it)

# --- Mechanic: the wrench is the business model ---
me = sheet_scaffold("Mechanic", "Farms lightly (hay only); income is repair contracts and a salvage restoration flip. The baler fix on day 4 opens the co-op repair pipeline. Wins on skill, not acreage.")
ev(me, 2, "Plant hay — East field", I=f"={cr('hay','B')}+{ar('fuel')}")
ev(me, 4, "Fix Hollis's baler (check) — co-op repair contract unlocks")
ev(me, 6, "Neighbor repair job", G=f"={ar('nb_pay')}", I=f"={ar('fuel')}")
ev(me, 7, "Co-op repair job 1/3", G=f"={ar('rj_pay')}", I=f"={ar('rj_parts')}+{ar('fuel')}")
ev(me, 8, "Hay cut 1 — East", D=f"={cr('hay','G')}*{cr('hay','H')}", I=f"={cr('hay','E')}")
ev(me, 9, "Co-op repair job 2/3", G=f"={ar('rj_pay')}", I=f"={ar('rj_parts')}+{ar('fuel')}")
ev(me, 10, "Buy salvage baler at Weaver's yard", I=f"={ar('flip_buy')}+{ar('fuel')}")
ev(me, 11, "Co-op repair job 3/3", G=f"={ar('rj_pay')}", I=f"={ar('rj_parts')}+{ar('fuel')}")
ev(me, 12, "Restoration parts for the flip", I=f"={ar('flip_parts')}")
ev(me, 15, "Hay cut 2 — East", D=f"={cr('hay','G')}*{cr('hay','H')}", I=f"={cr('hay','E')}")
ev(me, 18, "Neighbor repair job 2", G=f"={ar('nb_pay')}", I=f"={ar('fuel')}")
ev(me, 22, "Hay cut 3 — East", D=f"={cr('hay','G')}*{cr('hay','H')}", I=f"={cr('hay','E')}")
ev(me, 24, "Sell restored baler (flip complete)", G=f"={ar('flip_sale')}", I=f"={ar('fuel')}")
me_sum = summary(me)

# ---------------- Comparison ----------------
cp = wb.create_sheet("Comparison")
cp.column_dimensions["A"].width = 26
for c in "BCD":
    cp.column_dimensions[c].width = 18
put(cp, "A1", "THREE PATHS, SAME 30 DAYS", H1)
put(cp, "A2", "Phase 0 exit criterion: every background solvent, every strategy distinct.", Font(name="Arial", size=9, italic=True))
names = ["Old School", "IT Nephew", "Mechanic"]
sums = [os_sum, it_sum, me_sum]
for i, n in enumerate(names):
    put(cp, f"{get_column_letter(2+i)}4", n, H2)
    cp[f"{get_column_letter(2+i)}4"].fill = HDRFILL
metrics = [("End cash", 6, MONEY), ("End debt", 7, MONEY), ("Net worth delta", 8, MONEY),
           ("Top income source", 9, None),
           ("Spot sales", 0, MONEY), ("Negotiated (rep)", 1, MONEY), ("Timed (info)", 2, MONEY),
           ("Repair & flip", 3, MONEY), ("Eggs & favors", 4, MONEY)]
for j, (label, off, fmt) in enumerate(metrics):
    put(cp, f"A{5+j}", label, BOLD)
    for i, n in enumerate(names):
        put(cp, f"{get_column_letter(2+i)}{5+j}", f"='{n}'!B{sums[i]+off}", GREEN, fmt)
put(cp, "A16", "CHECKS", H2); cp["A16"].fill = HDRFILL
put(cp, "A17", "All solvent (end cash > 0)?", BOLD)
put(cp, "B17", '=IF(AND(B5>0,C5>0,D5>0),"PASS","FAIL")', BLACK)
put(cp, "A18", "All improved net worth?", BOLD)
put(cp, "B18", '=IF(AND(B7>0,C7>0,D7>0),"PASS","FAIL")', BLACK)
put(cp, "A19", "Top income sources all different?", BOLD)
put(cp, "B19", '=IF(AND(B8<>C8,B8<>D8,C8<>D8),"PASS","FAIL")', BLACK)
put(cp, "A20", "Balance fairness (max/min end cash ≤ 1.25)?", BOLD)
put(cp, "B20", '=IF(MAX(B5:D5)/MIN(B5:D5)<=1.25,"PASS","FAIL")', BLACK)

put(cp, "A23", "Daily cash curves", H2); cp["A23"].fill = HDRFILL
put(cp, "A24", "Day", BOLD)
for i, n in enumerate(names):
    put(cp, f"{get_column_letter(2+i)}24", n, BOLD)
for d in range(1, NDAYS + 1):
    r = 24 + d
    put(cp, f"A{r}", d, BLACK, "0")
    for i, n in enumerate(names):
        put(cp, f"{get_column_letter(2+i)}{r}", f"='{n}'!K{DAY0+d-1}", GREEN, MONEY)
chart = LineChart()
chart.title = "Cash by day — three backgrounds"
chart.height = 9
chart.width = 22
data = Reference(cp, min_col=2, max_col=4, min_row=24, max_row=54)
cats = Reference(cp, min_col=1, min_row=25, max_row=54)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
cp.add_chart(chart, "F4")

wb.save(r"C:\dev\Dirt_Money\design\economy_model.xlsx")
print("saved")
